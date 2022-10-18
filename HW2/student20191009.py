#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

#총점 계산
row_id = 1
total = []
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append([row_id, sum_v])
	row_id += 1

#내림차순 정렬
totalList = sorted(total, key = lambda x:x[1], reverse = True)
students = len(total)

#학점 범위 설정
AZ = int(students * 0.3)
AP = int(AZ * 0.5)
BZ = int(students * 0.7) - AZ
BP = int(BZ * 0.5)
CZ = students - AZ - BZ
CP = int(CZ * 0.5)

#순차적으로 성적에 따라 grade 설정
for i in range(AZ):
	ws.cell(row = totalList[i][0], column = 8).value = 'A0'
for i in range(AP):
	ws.cell(row = totalList[i][0], column = 8).value = 'A+'
for i in range(AZ, AZ + BZ):
	ws.cell(row = totalList[i][0], column = 8).value = 'B0'
for i in range(AZ, AZ + BP):
	ws.cell(row = totalList[i][0], column = 8).value = 'B+'
for i in range(AZ + BZ, AZ + BZ + CZ):
	ws.cell(row = totalList[i][0], column = 8).value = 'C0'
for i in range(AZ + BZ, AZ + BZ + CP):
	ws.cell(row = totalList[i][0], column = 8).value = 'C+'

wb.save("student.xlsx")

